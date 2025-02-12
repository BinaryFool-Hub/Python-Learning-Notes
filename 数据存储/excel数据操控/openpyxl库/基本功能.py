import openpyxl

workbook = openpyxl.load_workbook('测试.xlsx')  # 读取EXCEL文件变成对象
# workbook = openpyxl.Workbook()  # 新建EXCEL文件对象

sheet = workbook.active  # 选中活动工作表
# sheet = workbook.create_sheet('新工作表')  # 创建新的工作表对象
# sheet = workbook['Sheet']  # 选中指定的工作表

print(workbook.sheetnames)  # 获取工作簿中所有表的名字
print(sheet.max_row)  # 获取表格最大行
print(sheet.max_column)  # 获取表格最大列

workbook.save('测试.xlsx')
