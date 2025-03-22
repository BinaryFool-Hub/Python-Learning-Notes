import openpyxl

workbook = openpyxl.load_workbook('测试.xlsx')  # 读取文件构成对象
sheet = workbook.active  # 选中活动工作表

"""遍历所有数据"""
for row in sheet.iter_rows(values_only=True):
    # values_only=True 表示仅返回单元格的值，而不包含单元格对象（即不会返回 Cell 对象，而是直接返回对应的值）。
    # 如果不加需要在遍历row再使用.value来获取
    print(row)

"""读取单元格数据，无则返回None"""
cell_value = sheet['A3'].value
print(cell_value)
cell_value = sheet.cell(row=1, column=3).value
print(cell_value)

"""读取整行或整列的数据"""
row_data = [cell.value for cell in sheet[1]]  # 读取第一行数据
column_data = [cell.value for cell in sheet['A']]  # 读取第一列数据
print(row_data)
print(column_data)
